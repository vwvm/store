import re
from time import sleep

import openpyxl
import requests
import pyperclip
from PySide6.QtWidgets import QApplication, QFileDialog
from PySide6.QtCore import QObject, Signal, Slot, QThread, QRunnable, QThreadPool
from PySide6.QtQml import QQmlApplicationEngine
import sys
import os

from fontTools.varLib.instancer import setMvarDeltas
from openpyxl import load_workbook
from man_api import send_get_request, send_post_request
from qiniu_server import QiniuServer

qiniu_server = QiniuServer()


def download_video(url, save_folder, file_name=None):
    """
    下载视频方法
    @param url: 线上视频地址
    @param save_folder: 保存的文件夹路径
    @param file_name: 保存的文件名（可选）
    """
    try:
        # 如果文件名没有传入，则使用默认名称
        if file_name is None:
            file_name = "default_video_name.mp4"

        # 检查 save_folder 是否是一个文件夹路径，如果不是，则抛出异常
        if not os.path.isdir(save_folder):
            raise ValueError(f"提供的路径不是一个有效的文件夹: {save_folder}")

        # 拼接保存的完整路径
        save_path = os.path.join(save_folder, file_name)

        # 检查文件夹部分是否存在，不存在则创建
        if not os.path.exists(save_folder):
            os.makedirs(save_folder)

        # 发送请求获取视频内容，流式下载大文件
        response = requests.get(url, stream=True)
        response.raise_for_status()  # 检查请求是否成功

        # 打开文件并写入视频内容
        with open(save_path, 'wb') as video_file:
            for chunk in response.iter_content(chunk_size=1024 * 1024):  # 使用1MB块读取
                if chunk:
                    video_file.write(chunk)

        print(f"视频已保存到: {save_path}")

    except requests.exceptions.RequestException as e:
        print(f"视频下载失败: {e}")
    except ValueError as ve:
        print(f"输入错误: {ve}")


def get_excel_ce(file_path):
    """
    用来确定列中是否有数据
    :param file_path: Excel 文件的路径
    :return: 没有数据的行
    """
    # 加载 Excel 文件
    wb = openpyxl.load_workbook(file_path)
    # 选择第一个表格
    sheet = wb.worksheets[0]

    # 从 A3 开始检查
    row = 3
    while True:
        cell_value = sheet[f'A{row}'].value
        if cell_value is not None:
            print(f'A{row} 有数据: {cell_value}')
        else:
            print(f'A{row} 没有数据')
            break
        row += 1
    return row

def update_excel_cell(file_path: str, sheet_name: str, cell: str, value: str) -> None:
    """
    更新excel单元格
    @param file_path: 文件路径
    @param sheet_name: 表名称
    @param cell: 单元格位置
    @param value: 值
    @return:
    """
    # 使用openpyxl加载Excel文件
    workbook = load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' does not exist in the file.")
        return

    # 选择工作表并更新单元格内容
    sheet = workbook[sheet_name]
    sheet[cell] = value

    # 保存更改
    workbook.save(file_path)
    print(f"成功更新文件 '{file_path}' 中表 '{sheet_name}' 的单元格 '{cell}' 为 '{value}'！")


class WorkerSignals(QObject):
    finished = Signal()
    progress = Signal(str)
    setModelValue = Signal(int, str, str)


class DigitalHumanWorker(QRunnable):
    def __init__(self, model, access_token):
        super().__init__()
        self.model = model
        self.signals = WorkerSignals()
        self.access_token = access_token

    @Slot()
    def run(self):
        try:
            # 主要部分
            # print(self.model.get("outputPath2"))
            # download_video(
            #     "https://res.chanjing.cc/chanjing/dp/output/2024-10-21/1729502961055-7c90fe1c919aa3658b1fc53f72a5c64a-video.mp4",
            #     self.model.get("outputPath2"))
            # self.signals.setModelValue.emit(self.model.get("index"), "state_label", "生成数字人视频完成")
            # return

            # 获取文件路径
            filePath = self.model.get("filePath")
            if not filePath:
                self.signals.finished.emit()
                self.signals.setModelValue.emit(self.model.get("index"), "state_label", "请【选择文件】")
                return

            # 获得数字人名称
            digitalHumanName = self.model.get("digitalHumanName")
            if not digitalHumanName:
                self.signals.finished.emit()
                self.signals.setModelValue.emit(self.model.get("index"), "state_label", "请输入数字人名称")
                return
            # 获得文案
            digitalHumanText = self.model.get("scriptText")
            if not digitalHumanText:
                self.signals.finished.emit()
                self.signals.setModelValue.emit(self.model.get("index"), "state_label", "请输入文案")
                return
            # 处理掉文案的所有空字符
            digitalHumanText = re.sub(r"\s+", "", digitalHumanText)
            # 获得保存路径
            outputPath2 = self.model.get("outputPath2")
            if not outputPath2:
                self.signals.finished.emit()
                self.signals.setModelValue.emit(self.model.get("index"), "state_label", "请【设置视频保存的路径】")
                return

            # 获取文件路径上传七牛
            self.signals.setModelValue.emit(self.model.get("index"), "state_label", "上传视频中。。。")
            key = qiniu_server.file_upload(filePath)
            if not key:
                self.signals.finished.emit()
                self.signals.setModelValue.emit(self.model.get("index"), "state_label", "视频上传失败")
                return
            self.signals.setModelValue.emit(self.model.get("index"), "state_label", "上传视频完成,生成数字人中。")

            # 生成数字人
            data = {
                "name": digitalHumanName,
                "material_video": qiniu_server.get_object_url(key)
            }
            res = send_post_request(url="https://www.chanjing.cc/api/open/v1/create_customised_person",
                                    data=data,
                                    access_token=self.access_token)
            code = res.get("code")
            if code != 0:
                self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                                "创建数字人任务失败：" + res.get("msg"))
                return

            man_id = res.get("data")

            is_while = True
            while is_while:
                sleep(1)
                res = send_get_request(url="https://www.chanjing.cc/api/open/v1/customised_person",
                                       params={"id": man_id},
                                       access_token=self.access_token)
                self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                                "上传视频完成,生成数字人中。。" + str(res.get("data").get("progress")))
                if res.get("data").get("progress") == 100:
                    is_while = False

            self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                            "生成数字人完成。" + str(res.get("data").get("progress")))
            self.signals.setModelValue.emit(self.model.get("index"), "image_url", res.get("data").get("pic_url"))

            # 生成数字人视频
            audio_man_id = res.get("data").get("audio_man_id")
            video_width = res.get("data").get("width")
            video_height = res.get("data").get("height")
            data = {
                "person": {
                    "id": man_id,
                    "x": 0,
                    "y": 0,
                    "width": video_width,
                    "height": video_height
                },
                "audio": {
                    "tts": {
                        "text": [
                            digitalHumanText
                        ],
                        "speed": 20,
                        "audio_man": audio_man_id
                    },
                    "wav_url": "",
                    "type": "tts",
                    "volume": 100,
                    "language": "cn"
                },
                "bg_color": "#EDEDED",
                "screen_width": video_width,
                "screen_height": video_height
            }
            res = send_post_request(url="https://www.chanjing.cc/api/open/v1/create_video",
                                    data=data,
                                    access_token=self.access_token)
            code = res.get("code")
            if code != 0:
                self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                                "创建数字人视频任务失败" + res.get("message"))
                return
            video_id = res.get("data")

            is_while = True
            while is_while:
                sleep(1)
                res = send_get_request(url="https://www.chanjing.cc/api/open/v1/video",
                                       params={"id": video_id},
                                       access_token=self.access_token)
                self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                                "生成数字人视频中。。" + str(res.get("data").get("progress")))
                if res.get("data").get("progress") == 100:
                    is_while = False
            self.signals.setModelValue.emit(self.model.get("index"), "state_label",
                                            "生成数字人视频中。" + str(res.get("data").get("progress")))
            self.signals.setModelValue.emit(self.model.get("index"), "outputPath1", res.get("data").get("video_url"))

            # 保存线上视频
            download_video(res.get("data").get("video_url"), outputPath2, digitalHumanName)
            self.signals.setModelValue.emit(self.model.get("index"), "state_label", "生成数字人视频完成")

            # 获得数字人视频id和音频id加上文案生成数字人视频
        except Exception as e:
            self.signals.setModelValue.emit(self.model.get("index"), "state_label", "程序错误" + str(e))
        finally:
            # 完成后发出信号
            self.signals.finished.emit()


class Backend(QObject):
    def __init__(self, engine):
        super().__init__()
        self.thread_pool = QThreadPool()  # 创建线程池
        self.access_token = "AkA2MYVdpVT01HraCo/VMPngAiKndMmV8qzdAo7JPSpsp/FsvCFSu+RchrqKH7cR"
        self.engine = engine
        self.file_path = ""

    # 定义信号，用于将选择的文件路径发送回 QML，携带索引和文件路径
    filePathSelected = Signal(int, str)
    savePathSelected = Signal(int, str)
    # 切换页面信号
    switchPage = Signal(str)  # 用于通知 QML 切换页面
    # 设置model的值
    setModelValue = Signal(int, str, str)

    @Slot(str)
    def on_progress(self, message):
        # 处理进度消息，例如更新 UI 或打印日志
        print(message)

    @Slot()
    def on_worker_finished(self):
        # 处理 worker 完成的逻辑，例如通知 UI 或进行其他操作
        print("Worker has finished.")

    @Slot()
    def on_set_model_value(self, index, key, value):
        self.setModelValue.emit(index, key, value)

    @Slot(int)
    def open_file_dialog(self, index):
        # 打开文件选择对话框
        file_path, _ = QFileDialog.getOpenFileName(None, "选择文件", "", "All Files (*)")
        if file_path:
            # 发射信号，将索引和选择的文件路径发送回 QML 界面
            self.filePathSelected.emit(index, file_path)

    @Slot(str)
    def copy_to_clipboard(self, text):
        # 将路径复制到剪切板
        pyperclip.copy(text)
        self.setModelValue.emit(0, "state_label", "数字人生成完成：复制成功")

    @Slot(int)
    def set_save_file_path(self, index):
        # 打开文件夹选择对话框
        folder_path = QFileDialog.getExistingDirectory(None, "选择文件夹", "")

        if folder_path:
            # 使用 os.path.normpath() 将路径标准化，确保路径适合当前的操作系统
            folder_path = os.path.normpath(folder_path)

            # 发射信号，将索引和选择的文件夹路径发送回 QML 界面
            self.savePathSelected.emit(index, folder_path)

    @Slot(dict)
    def generate_digital_human_video(self, model):
        """
        多线程生成数值人视频任务
        @param model:
        """
        # 创建 worker 对象并添加到线程池中
        worker = DigitalHumanWorker(model, self.access_token)
        worker.signals.progress.connect(self.on_progress)
        worker.signals.finished.connect(self.on_worker_finished)
        worker.signals.setModelValue.connect(self.on_set_model_value)

        # 将任务交给线程池执行
        self.thread_pool.start(worker)

    # 用来打印信息
    @Slot(str)
    def slot_print(self, msg):
        print(msg)

    @Slot(str)
    def verify_input(self, access_token):
        """
        验证access_token
        @param access_token:
        """

        res = send_get_request(url="https://www.chanjing.cc/api/open/v1/user_info", access_token=access_token)
        if res.get("code") == 0:
            # 验证通过
            self.access_token = access_token
            self.switchPage.emit("view.qml")  # 验证通过后切换page
            pass
        else:
            # 验证不通过
            pass


def start_window():
    app = QApplication(sys.argv)
    engine = QQmlApplicationEngine()

    # 创建 Backend 实例，并传递 engine
    backend = Backend(engine)
    engine.rootContext().setContextProperty("backend", backend)

    # 加载 QML 文件
    qml_file_path = "qml/main.qml"  # 加载qml
    engine.load(qml_file_path)

    if not engine.rootObjects():
        sys.exit(-1)

    sys.exit(app.exec())
    pass


if __name__ == '__main__':
    # access_key = "zknxbe6OotPFY3ZfbcdCCjRY_ftPAYx-A55PECW9"
    # secret_key = ""
    # q = Auth(access_key, secret_key)

    # 更新Excel文件中某个单元格

    # download_video(
    #     "https://res.chanjing.cc/chanjing/dp/output/2024-10-21/1729502961055-7c90fe1c919aa3658b1fc53f72a5c64a-video.mp4",
    #     r"F:\Users\BlackBox\project\synthetic\store\store_python\script\zotci_digital_man.mp4"
    # )
    start_window()

    # update_excel_cell('数字人交付表格.xlsx', 'Sheet1', 'b3', '更新内容')
